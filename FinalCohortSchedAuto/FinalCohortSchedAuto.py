from openpyxl import load_workbook
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils import column_index_from_string
from openpyxl.styles import NamedStyle
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Font, Border, Fill, Alignment, Protection
import traceback
from datetime import datetime, timedelta
from tkinter import Toplevel, PhotoImage
from tkinter import filedialog
from datetime import datetime
from openpyxl.styles.numbers import FORMAT_DATE_DDMMYY
from PIL import Image, ImageTk
import configparser
import shutil
import tkinter as tk
from tkinter import ttk
import os

class App:
    def __init__(self, root):
        self.root=root
        self.source_file_path = ""
        self.source_file_name = ""
        self.target_file_path = ""
        self.target_file_name = ""
        self.error_list = []
        self.complete_list = []
        self.folder_mode_var = tk.BooleanVar()  # Variable for "Folder Mode" checkbox
        self.create_backup_var = tk.BooleanVar()
        self.progbarMultiplier = 0 #Multiplier for the progress bar
        self.progLength = 200
        self.backupFlag = 1
#/--------------------------------GUI------------------------------/
        #setting title
        root.title("Tammy's Cohort Sched Tool")
        #setting window size
        root.geometry("450x180")
        root.resizable(width=False, height=False)

        #Define mode flag [single or folder]
        self.modeFlag = 0 #0 - singleSubmit 1 - folderSubmit

        # Source file label
        self.source_label = tk.Label(root, text="Calendar:")
        self.source_label.place(x=30,y=30)

        self.source_file_label = tk.Label(root, text='No File Selected')
        self.source_file_label.place(x=150,y=30)

        # Source file browse button
        self.source_browse_button = tk.Button(root, text="Browse", command=self.browse_source, bg="#8AC6D1", fg="white", width=10, height=1, relief='raised')
        self.source_browse_button.place(x=330,y=30)
        
        # Target file label
        self.target_label = tk.Label(root, text="Cohort Schedule:")
        self.target_label.place(x=30,y=70)

        self.target_file_label = tk.Label(root, text='No File Selected')
        self.target_file_label.place(x=150,y=70)

        # Target file browse button
        self.target_browse_button = tk.Button(root, text="Browse", command=self.browse_target, bg="#8AC6D1", fg="white", width=10, height=1, relief='raised')
        self.target_browse_button.place(x=330,y=70)

        # Submit button. Initially blue for single mode.
        self.submit_button = tk.Button(root, text="File Submit", command=self.singleSubmit, bg="#8AC6D1", fg="white", width=20, height=2)
        self.submit_button.place(x=150,y=115)
        # Create the second button (red button for Folder mode)
        self.second_button = tk.Button(root, text="Folder Submit", command=self.folderSubmit, bg="#FFB6C1", fg="white", width=20, height=2)
        # Initially hide the red button
        self.second_button.place_forget()

        # Load the original image
        original_image = Image.open("settings_cog.gif")

        # Define the desired size
        desired_size = (30, 30)  # Replace width and height with your desired values

        # Resize the image
        resized_image = original_image.resize(desired_size, Image.LANCZOS)

        # Convert the resized image to PhotoImage
        settings_icon = ImageTk.PhotoImage(resized_image)

        # Create the button with the resized image
        settings_button = tk.Button(root, image=settings_icon, command=self.open_settings, bg="#8AC6D1", relief='raised')
        settings_button.place(x=30, y=120)
        settings_button.image = settings_icon  # keep a reference of the image

        #Status notification
        self.status_label = tk.Label(root, text="")
        self.status_label.place(x=330, y=125)
        
        # Load the checkbox state
        self.load_checkbox_state()
    
    def create_progressbar(self, multiplier):
        if (self.modeFlag):
            self.progbarMultiplier = multiplier
        # Create the progress bar
        self.progress = ttk.Progressbar(self.root, length=200, mode='determinate')
        self.progress.pack()

    def updateProgress(self):
        if (self.modeFlag):
            self.progress['value'] = self.progress['value'] + (.1/self.progbarMultiplier)
        else:
            self.progress['value'] += .1 
        

        root.update_idletasks()

    def switch_button(self):
        if self.folder_mode_var.get():
            # Hide the blue button and show the red button
            self.submit_button.place_forget()
            self.second_button.place(x=150, y=115)
        
            # Change the text of the source_label and the color of the browse button
            self.source_label.config(text="Folder:")
            self.source_browse_button.config(bg="#FFB6C1")
        else:
            # Hide the red button and show the blue button
            self.second_button.place_forget()
            self.submit_button.place(x=150, y=115)

            # Change the text of the source_label and the color of the browse button back to original
            self.source_label.config(text="Calendar:")
            self.source_browse_button.config(bg="#8AC6D1")

    def save_checkbox_state(self):
        config = configparser.ConfigParser()
        config.read('settings.ini')  # Read the existing configuration file

        if 'Settings' not in config:
            config['Settings'] = {}

        config['Settings']['CreateBackup'] = str(self.create_backup_var.get())  # Save the checkbox state
        config['Settings']['FolderMode'] = str(self.folder_mode_var.get()) 

        with open('settings.ini', 'w') as configfile:
            config.write(configfile)  # Write the updated configuration to the file

    def load_checkbox_state(self):
        config = configparser.ConfigParser()
        config.read('settings.ini')  # Read the configuration file

        if 'Settings' in config and 'CreateBackup' in config['Settings']:
            self.create_backup_var.set(config.getboolean('Settings', 'CreateBackup'))  # Load the checkbox state
        if 'Settings' in config and 'FolderMode' in config['Settings']:
            self.folder_mode_var.set(config.getboolean('Settings', 'FolderMode'))  # Load the checkbox state
    
    def load_checkbox_state(self):
        config = configparser.ConfigParser()
        config.read('settings.ini')  # Read the configuration file

        if 'Settings' in config and 'CreateBackup' in config['Settings']:
            self.create_backup_var.set(config.getboolean('Settings', 'CreateBackup'))  # Load the checkbox state
        if 'Settings' in config and 'FolderMode' in config['Settings']:
            self.folder_mode_var.set(config.getboolean('Settings', 'FolderMode'))  # Load the checkbox state
    
        self.switch_button()  # Switch the buttons according to the checkbox state

    def open_settings(self):
        self.load_checkbox_state()
        self.switch_button()  # Switch the buttons according to the checkbox state

        settings_window = Toplevel(root)  # Use the 'root' directly
        settings_window.title("Settings")
        settings_window.geometry("150x150")
        settings_window.resizable(width=False, height=False)

        # Backup Checkbox
        create_backup_checkbox = tk.Checkbutton(settings_window, text="Create Backup", variable=self.create_backup_var)
        create_backup_checkbox.pack()
        create_backup_checkbox.place(x=30, y=30)

        # Folder Mode Checkbox
        folder_mode_checkbox = tk.Checkbutton(settings_window, text="Folder Mode", variable=self.folder_mode_var, command=self.switch_button)
        folder_mode_checkbox.pack()
        folder_mode_checkbox.place(x=30, y=60)

        self.load_checkbox_state()

        def apply_settings():
            self.save_checkbox_state()
            settings_window.destroy()

        # Apply Button
        apply_button = tk.Button(settings_window, text="Apply", command=apply_settings)
        apply_button.pack()
        apply_button.place(x=30, y=90)
        # You can add more elements to the settings window here

    def browse_source(self):
        if self.folder_mode_var.get():
            # Folder mode is selected, let user choose a directory
            source_folder = filedialog.askdirectory(title = "Select a Folder")
            self.source_file_path = os.path.realpath(source_folder)
            self.source_file_label.config(text=os.path.basename(source_folder))
        else:
            # Folder mode is not selected, let user choose a file
            source_file = filedialog.askopenfilename(title = "Select a File", filetypes = (("Excel files", "*.xlsx"),("Excel files", "*.xls")))
            self.source_file_path = os.path.realpath(source_file)
            self.source_file_label.config(text=os.path.basename(source_file))
            self.source_file_name = os.path.basename(source_file)
        self.status_label.config(text="")

    def browse_target(self):
        target_file = filedialog.askopenfilename(title = "Select a File", filetypes = (("Excel files", "*.xlsx"),("Excel files", "*.xls")))
        self.target_file_path = os.path.realpath(target_file)
        self.target_file_label.config(text=os.path.basename(target_file))
        self.target_file_name = os.path.basename(target_file)
        self.status_label.config(text="")
   
    def show_error_window(self, error_report):
        # Create a new window
        error_window = tk.Toplevel()
        error_window.geometry("500x300")  # you can adjust the window size
        error_window.resizable(width=False, height=False)

        # Add a title to the window
        error_window.title("Error Report")

        # Create a frame to add padding
        frame = tk.Frame(error_window, bg="lightgray", padx=20, pady=20)
        frame.pack(fill='both', expand=True)

        # Create a scrollable text widget
        text_widget = tk.Text(frame, wrap='word', height=10, width=50, bg="lightgray", fg="red", font=("Arial", 10))  # you can adjust height and width
        text_widget.pack(side='left', fill='both', expand=True)

        # Add a scrollbar widget
        scrollbar = tk.Scrollbar(frame)
        scrollbar.pack(side='right', fill='y')

        # Configure the scrollbar to scroll the text widget
        text_widget.configure(yscrollcommand=scrollbar.set)
        scrollbar.configure(command=text_widget.yview)

        # Insert the error report into the text widget
        text_widget.insert('1.0', error_report)

        # Disable editing
        text_widget.configure(state='disabled')

        # Create an OK button that will destroy the window when clicked
        ok_button = tk.Button(frame, text="OK", command=error_window.destroy, bg="lightgray", fg="black", relief='raised', font=("Arial", 12), padx=20, pady=10)
        ok_button.pack(pady=10)

        # This line is needed to handle window events
        error_window.mainloop()


#/--------------------------------Submit Algorithm------------------------------/
    #Submit button function. Will attempt to load CST.xlsx file first, then impor the data
    #from the calendar into the CST.xlsx file.
    def folderSubmit(self):
        multiplier = 0
        self.modeFlag = 1 #Set the mode to folderSubmit
        try:
            # Save the original directory path
            source_folder_path = self.source_file_path

            # Get list of all files in the selected directory
            files_in_directory = os.listdir(self.source_file_path)

            # Filter the list to include only .xlsx and .xls files
            excel_files = [file for file in files_in_directory if file.endswith((".xlsx", ".xls"))]
            
            for file in excel_files:
                multiplier = multiplier + 1
            self.create_progressbar(multiplier)

            # Loop over each file and call singleSubmit on each file
            for file in excel_files:
                self.source_file_path = os.path.join(source_folder_path, file)  # use source_folder_path
                self.source_file_name = os.path.basename(file)
                try:
                    self.singleSubmit()  # call singleSubmit on each file
                except Exception as e:
                        print(e)
                        self.status_label.config(text="Error Added")
                        self.error_list.append(self.source_file_name)
                        print(self.source_file_name)
                else:
                        self.complete_list.append(self.source_file_name)

            self.source_file_path = source_folder_path  # reset source_file_path
            self.progress['value'] += 200 
             # once everything else is done, display the error report
            if self.error_list:
                error_report = "Manual Addition Required:\n" + "\n".join(self.error_list)
                print(error_report)
            if self.complete_list:
                complete_report = "Completed:\n" + "\n".join(self.complete_list)
                print(complete_report)
                # display `error_report` in a window python setup.py py2exe

            self.show_error_window(error_report)
            self.show_error_window(complete_report)
        except Exception as e:
            print(e)
            traceback.print_exc()

    def singleSubmit(self):
        if (self.modeFlag == 0): #if not in folder mode
            self.create_progressbar(1) #create progress bar with dimension 1

            # here you have access to the source and target files
            wb1 = load_workbook(self.target_file_path)
            wb2 = load_workbook(self.source_file_path)

            wb1.save(self.target_file_path)
            self.status_label.config(text="Loading..")

        # here you have access to the source and target files
        source_file_name = self.source_file_label['text']
        target_file_name = self.target_file_label['text']

        source_file_path = self.source_file_path
        target_file_path = self.target_file_path

        # Prepare Cohort Sched excel workbook and sheet to be added
        wb1 = load_workbook(self.target_file_path)
        wb2 = load_workbook(self.source_file_path)
       
        while True:
            if self.backupFlag == 0:
                break
                # Create backup if the user has selected that option
            if self.create_backup_var.get():
                backup_folder = tk.filedialog.askdirectory()
                if backup_folder:  # Make sure the user selected a directory
                    backup_path = os.path.join(backup_folder, os.path.basename(target_file_path))
                    shutil.copy2(target_file_path, backup_path)
            self.backupFlag = 0


        # Obtain primary worksheet from source
        wsC = wb2.worksheets[0]

        # Split name of wb2 to isolate name of program
        targetWS = self.source_file_name.split()
        ws = wb1[targetWS[0]]
       
        base_cell= wb1[targetWS[0]]['A3'] #obtain the base style of the cells
        # Create a new style from the base cell's style
        base_font = Font(name=base_cell.font.name, bold=base_cell.font.bold, italic=base_cell.font.italic)
        base_border = Border(left=base_cell.border.left, right=base_cell.border.right, top=base_cell.border.top, bottom=base_cell.border.bottom)
        base_alignment = Alignment(horizontal=base_cell.alignment.horizontal, vertical=base_cell.alignment.vertical)
        base_protection = Protection(locked=base_cell.protection.locked, hidden=base_cell.protection.hidden)

        #Find number of rows to be copied and inserted. compare val of a and b cells, 
        #increase b by one (one row down) and repeat until section is over. a starts at 3 to make up for header
        #b starts at i (4) right under 3, I is incremented each comparison.
        i = 4
        a = ws.cell(row=3, column=1)
        b = ws.cell(row=i, column=1)
        counter = 2

        while (a.value == b.value):
            counter += 1
            i += 1
            b = ws.cell(row=i, column=1)  
            
       
        # Read existing rows from the target worksheet and store them in memory
        existing_rows = []
        merged_cells_info = []
        hidden_rows = []
        for row in ws.iter_rows(min_row=2, max_col=12):  # Define the range of columns here
            self.updateProgress() #######################################################################################
            row_data = {}
            for cell in row:
                if isinstance(cell, MergedCell):  # Skip merged cells
                    continue
                row_data[cell.column_letter] = cell.value
                if ws.row_dimensions[cell.row].hidden:
                    row_data["hidden"] = True
                    hidden_rows.append(cell.row)   # Keep track of hidden rows
            existing_rows.append(row_data)
            # Store merged cells' information (cell range and data)
            for merged_cell_range in ws.merged_cells.ranges:
                top_left_cell = ws[merged_cell_range.start_cell.coordinate]
                merged_cells_info.append((merged_cell_range, top_left_cell.value))

        # Add new rows at the top of the target worksheet
        ws.insert_rows(2, counter)

        for row_index, row_data in enumerate(existing_rows, start=counter + 2):
            for col, value in row_data.items():
                if col != "hidden":
                    cell = ws.cell(row=row_index, column=column_index_from_string(col)) 
                    cell.value = value
            if "hidden" in row_data and row_data["hidden"]:
                ws.row_dimensions[row_index].hidden = True


        # Reapply merged cell ranges and restore their data
        print("Before loop, merged_cells_info:", len(merged_cells_info))

        merged_cells_info_copy = merged_cells_info.copy()
        merged_cells_info_copy = list(set(merged_cells_info_copy)) 
        while merged_cells_info_copy:
            self.updateProgress() 
            merged_cell_range, cell_value = merged_cells_info_copy[0]
            print("Inside loop, merged_cell_range:", merged_cell_range)
            ws.merge_cells(start_row=merged_cell_range.bounds[1] + counter, start_column=merged_cell_range.bounds[0],
                            end_row=merged_cell_range.bounds[3] + counter, end_column=merged_cell_range.bounds[2])
            top_left_cell = ws.cell(row=merged_cell_range.bounds[1] + counter, column=merged_cell_range.bounds[0])
            top_left_cell.value = cell_value
            merged_cells_info_copy.pop(0)
            # Remove the processed merged cell range from the list
            for i, (merged_cell_range_copy, _) in enumerate(merged_cells_info_copy):
                if merged_cell_range_copy == merged_cell_range:
                    del merged_cells_info_copy[i]
                    break
            print("merged_cells_info_copy:", merged_cells_info_copy)

        print("After loop, merged_cells_info:", len(merged_cells_info_copy))

        # After reapplying merged cells, adjust hidden rows
        new_hidden_rows = [row + counter for row in hidden_rows]

        # Unhide original rows and hide new ones
        for row in hidden_rows:
            ws.row_dimensions[row].hidden = False

        for row in new_hidden_rows:
            ws.row_dimensions[row].hidden = True


        # Title for new calendar added is program name + section 
        ws.cell(row=2, column=1).value = targetWS[0] + ' ' + targetWS[1]  

        #obtain section from copied ws and paste under 'program'
        for row in ws.iter_rows(min_row=3, max_col=1, max_row=counter+1):
            for cell in row:
                cell.value = targetWS[1]

        #obtain day and paste under Day section
        for row in ws.iter_rows(min_row=3, min_col=4, max_col=4, max_row=counter+1):
            for cell in row:
                cell.value = targetWS[2]
                if ".xlsx" in cell.value:
                    cell.value = cell.value.replace(".xlsx", "") 

        def iterate():
            for row in wsC.iter_rows():
                self.updateProgress() 
                for cell in row:
                    if cell in cellChecker:  # make sure not to repeat the same cell
                        continue

                    if cell.value is not None and not (isinstance(cell.value, str) and cell.value.strip() == ""):
                        a = str(cell.value).split()
                        if a[0] == 'Term':
                            global term
                            cellChecker.append(cell)
                            term = a[2]
                        # Find a course row based on a number existing on column B (the credits cell)
                        if cell.column_letter == 'B' and str(cell.value).isnumeric():
                            # Check if adjacent cells are not None before adding cell to the list
                            adjacent_cell1 = cell.offset(0, -1).value
                            adjacent_cell2 = cell.offset(0, 1).value
                            if adjacent_cell1 is not None and adjacent_cell2 is not None:
                                cellChecker.append(cell)
                                print(f"Returning cell {cell.coordinate}: {cell.value}")  # Add this line for debugging
                                return cell
            return None



    
        rowCounter=0
        row = 3
        col = 2
        currentRow = 0
        currentCol = 0
        global cellChecker
        cellChecker = []

        #row, col determine coordinates in CST doc. currentRow currentCol determine cooridnates on schedule to be added.
        #Counter is subtracted 1 to account for op header rows 1 and 2 
        while(rowCounter != counter-1):
            #Find cells with a 'credit' value, meaning rows with course listing on them. Obtain global 'term' for current term as well.
            self.updateProgress() 
            currentCell = iterate()
            
            if currentCell is None:
            # Handle the case when no matching cell is found
                print("No more matching cells found.")
                break

            #Write down course name and number
            ws.cell(row, col).value = str(currentCell.offset(0, -1).value).split()[0] + " " + str(currentCell.offset(0, -1).value).split()[1].replace(':', '')
            cell = ws.cell(row, col)
            cell = ws.cell(row=row, column=col)
            cell.font = base_font
            cell.border = base_border
            cell.alignment = base_alignment
            cell.protection = base_protection
            cell.number_format = base_cell.number_format


            #Write down term
            ws.cell(row, col+1).value = term
            cell = ws.cell(row, col+1)
            cell.font = base_font
            cell.border = base_border
            cell.alignment = base_alignment
            cell.protection = base_protection
            cell.number_format = base_cell.number_format

            #write down weeks
            # Write down weeks
            ws.cell(row, col+3).value = str(currentCell.offset(0, 2).value)
            cell = ws.cell(row, col+3)
            cell.font = base_font
            cell.border = base_border
            cell.alignment = base_alignment
            cell.protection = base_protection
            cell.number_format = base_cell.number_format

            # Write down dates Please note, datetime.strptime() function will throw an error if the date string is not in the expected format. You may need to handle this exception, especially if there's a chance that the date strings could be in a different format.
            # Write down dates
            ws.cell(row, col+5).value = str(currentCell.offset(0, 1).value).split()[0]
            cell = ws.cell(row, col+5)
            cell.font = base_font
            cell.border = base_border
            cell.alignment = base_alignment
            cell.protection = base_protection
            cell.number_format = base_cell.number_format
           
            ws.cell(row, col+6).value = str(currentCell.offset(0, 1).value).split()[2]
            cell = ws.cell(row, col+6)
            cell.font = base_font
            cell.border = base_border
            cell.alignment = base_alignment
            cell.protection = base_protection
            cell.number_format = base_cell.number_format

            #Move to next row for next iteration
            row += 1
            rowCounter += 1
          
                #show complete flag and save.
        try:
            wb1.save(self.target_file_path)
            self.status_label.config(text="           ")
            self.status_label.config(text="Complete!")
        except: 
            self.status_label.config(text="Error: Close Excel Files")


root = tk.Tk()
app = App(root)
root.mainloop()


